import os
import openpyxl

from PyQt5.QtGui import QIcon, QColor
from PyQt5.QtWidgets import QAction, QMessageBox
from PyQt5.QtCore import QVariant

from qgis.core import (
    QgsProject,
    QgsVectorLayer,
    QgsFeature,
    QgsGeometry,
    QgsPointXY,
    QgsField,
    QgsGraduatedSymbolRenderer,
    QgsClassificationEqualInterval,
    QgsSymbol,
    QgsPalLayerSettings,
    QgsVectorLayerSimpleLabeling,
    QgsVectorFileWriter,
    QgsDistanceArea,
    Qgis
)
class FlowMapGenerator:
    def __init__(self, iface):
        self.iface = iface
        self.plugin_dir = os.path.dirname(__file__)
        self.menu = "&Jeodezik Akış Haritası"

    def initGui(self):
        icon_path = os.path.join(self.plugin_dir, "icon.png")
        self.action = QAction(
            QIcon(icon_path),
            "Jeodezik Akış Haritası Oluştur",
            self.iface.mainWindow()
        )
        self.action.triggered.connect(self.run)
        self.iface.addToolBarIcon(self.action)
        self.iface.addPluginToMenu(self.menu, self.action)

    def unload(self):
        self.iface.removePluginMenu(self.menu, self.action)
        self.iface.removeToolBarIcon(self.action)

    def run(self):
        from .flowmap_dialog import FlowMapGeneratorDialog
        dlg = FlowMapGeneratorDialog()
        dlg.signal_paths_selected.connect(self.run_generation)
        dlg.exec_()

    def run_generation(self, airport_path, routes_path):
        try:
            airport_data = {}

            wb_air = openpyxl.load_workbook(airport_path)
            sh_air = wb_air.active

            for row in sh_air.iter_rows(min_row=2):
                aid, lat, lon = row[0].value, row[1].value, row[2].value
                if aid is not None and lat is not None and lon is not None:
                    airport_data[str(aid)] = QgsPointXY(float(lon), float(lat))

            if not airport_data:
                raise Exception("Havalimanı verisi bulunamadı")

            line_layer = QgsVectorLayer(
                "LineString?crs=EPSG:4326",
                "Jeodezik_Akimlar",
                "memory"
            )
            line_pr = line_layer.dataProvider()
            line_pr.addAttributes([
                QgsField("Origin", QVariant.String),
                QgsField("Dest", QVariant.String),
                QgsField("Flow", QVariant.Int),
                QgsField("Dist_km", QVariant.Double)
            ])
            line_layer.updateFields()

            point_layer = QgsVectorLayer(
                "Point?crs=EPSG:4326",
                "Havalimanlari",
                "memory"
            )
            point_pr = point_layer.dataProvider()
            point_pr.addAttributes([
                QgsField("IATA", QVariant.String)
            ])
            point_layer.updateFields()

            
            da = QgsDistanceArea()
            da.setEllipsoid("WGS84")
            da.setSourceCrs(
                line_layer.crs(),
                QgsProject.instance().transformContext()
            )
            wb_rot = openpyxl.load_workbook(routes_path)
            sh_rot = wb_rot.active

            lines = []
            points = []
            active_airports = set()

            for row in sh_rot.iter_rows(min_row=2):
                o_id = str(row[2].value)
                d_id = str(row[4].value)

                if o_id in airport_data and d_id in airport_data:
                    p1 = airport_data[o_id]
                    p2 = airport_data[d_id]

                   
                    distance_m = da.measureLine(p1, p2)
                    distance_km = distance_m / 1000.0

                    
                    segment_length = max(50000, distance_m / 20)

                    geodesic_result = da.geodesicLine(
                        p1,
                        p2,
                        segment_length
                    )

                    
                    if geodesic_result and isinstance(geodesic_result[0], list):
                        geodesic_points = geodesic_result[0]
                    else:
                        geodesic_points = geodesic_result

                    geom = QgsGeometry.fromPolylineXY(geodesic_points)

                    
                    flow_value = int(distance_km * 10)

                    f = QgsFeature(line_layer.fields())
                    f.setGeometry(geom)
                    f.setAttributes([
                        o_id,
                        d_id,
                        flow_value,
                        round(distance_km, 2)
                    ])

                    lines.append(f)
                    active_airports.update([o_id, d_id])

            
            for aid in active_airports:
                f = QgsFeature(point_layer.fields())
                f.setGeometry(QgsGeometry.fromPointXY(airport_data[aid]))
                f.setAttributes([aid])
                points.append(f)

            line_pr.addFeatures(lines)
            point_pr.addFeatures(points)

           
            self.apply_styles(line_layer, point_layer)

            save_path = os.path.join(
                os.path.expanduser("~"),
                "Documents",
                "akim_analizi.geojson"
            )

            options = QgsVectorFileWriter.SaveVectorOptions()
            options.driverName = "GeoJSON"
            options.fileEncoding = "UTF-8"

            result = QgsVectorFileWriter.writeAsVectorFormatV2(
                line_layer,
                save_path,
                QgsProject.instance().transformContext(),
                options
            )

            if isinstance(result, tuple):
                error = result[0]
            else:
                error = result

            if error != QgsVectorFileWriter.NoError:
                raise Exception("GeoJSON yazılamadı")

          
            QgsProject.instance().addMapLayer(line_layer)
            QgsProject.instance().addMapLayer(point_layer)

            self.iface.mapCanvas().setExtent(line_layer.extent())
            self.iface.mapCanvas().refresh()

            self.iface.messageBar().pushMessage(
                "Başarılı",
                "Jeodezik akış haritası (akış=mesafe türevli) oluşturuldu",
                level=Qgis.Success
            )

        except Exception as e:
            QMessageBox.critical(
                self.iface.mainWindow(),
                "Kritik Hata",
                str(e)
            )
        
    def apply_styles(self, l_layer, p_layer):
        renderer = QgsGraduatedSymbolRenderer("Flow")
        renderer.setClassificationMethod(QgsClassificationEqualInterval())
        renderer.updateClasses(l_layer, 5)

        for i, r in enumerate(renderer.ranges()):
            symbol = QgsSymbol.defaultSymbol(l_layer.geometryType())
            symbol.setWidth(0.8 + i * 0.7)
            symbol.setColor(QColor("#e31a1c"))
            r.setSymbol(symbol)

        l_layer.setRenderer(renderer)

        l_layer.setMapTipTemplate("""
            <h3>Jeodezik Rota</h3>
            <b>Başlangıç:</b> [% "Origin" %]<br>
            <b>Bitiş:</b> [% "Dest" %]<br>
			<br>
            <b>Jeodezik Mesafe:</b> [% format_number("Dist_km", 2) %] km<br>
			<b>Akış Büyüklüğü:</b> [% "Flow" %]<br>
			<b>Mesafe:</b> [% round($length/1000, 2) %] KM
            """)

        settings = QgsPalLayerSettings()
        settings.fieldName = "IATA"
        settings.enabled = True

        p_layer.setLabeling(QgsVectorLayerSimpleLabeling(settings))
        p_layer.setLabelsEnabled(True)
